import openpyxl
import re
import json
import os

def extract_url(text):
    if not text:
        return None, None
    text = str(text).strip()
    
    # 1. Match http/https URL
    match = re.search(r'https?://[^\s\"\'\>]+', text)
    if match:
        url = match.group(0).rstrip('.,;')
        return url, text
        
    # 2. Match www. domain
    match_www = re.search(r'www\.[a-zA-Z0-9\-\.]+\.[a-zA-Z]{2,}(?:/[^\s\"\'\>]*)?', text, re.IGNORECASE)
    if match_www:
        url = 'https://' + match_www.group(0).rstrip('.,;')
        return url, text
        
    # 3. Match known TLD domains (sportstimingsolutions.in, etc.)
    match_domain = re.search(r'[a-zA-Z0-9\-\.]+\.(com|in|org|net|co\.in|edu|gov)(?:/[^\s\"\'\>]*)?', text, re.IGNORECASE)
    if match_domain:
        url = 'https://' + match_domain.group(0).rstrip('.,;')
        return url, text
        
    return None, text

def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    f1 = os.path.join(base_dir, 'verified-records.xlsx')
    f2 = os.path.join(base_dir, 'verified-records-links.xlsx')
    
    print('Loading verified-records.xlsx...')
    wb1 = openpyxl.load_workbook(f1, data_only=True)
    r1 = list(wb1.active.iter_rows(values_only=True))
    h1 = r1[0]
    
    print('Loading verified-records-links.xlsx...')
    wb2 = openpyxl.load_workbook(f2, data_only=True)
    r2 = list(wb2.active.iter_rows(values_only=True))
    h2 = r2[0]
    
    runners = []
    mismatch_count = 0
    clean_url_count = 0
    cert_count = 0
    
    for i in range(1, len(r1)):
        row1 = dict(zip(h1, r1[i]))
        row2 = dict(zip(h2, r2[i]))
        
        # Clean URL extraction
        res_link_raw = row2.get('result_link')
        clean_url, raw_text = extract_url(res_link_raw)
        if clean_url:
            clean_url_count += 1
            
        # Certificate file processing
        c_file = row2.get('certificate_file')
        cert_path = None
        cert_type = None
        if c_file and str(c_file).startswith('certificates/'):
            cert_path = str(c_file).strip()
            ext = os.path.splitext(cert_path)[1].lower().lstrip('.')
            cert_type = ext if ext else 'unknown'
            cert_count += 1
            
        req_lineup = str(row1.get('requested_lineup_section')).strip() if row1.get('requested_lineup_section') else None
        exp_raw = row1.get('expected_lineup_section')

        if exp_raw:
            exp_lineup = str(exp_raw).strip()
        elif str(row1.get('verification_status')).strip() == 'Verified':
            exp_lineup = req_lineup
        else:
            exp_lineup = 'C'
        
        is_mismatch = (req_lineup != exp_lineup)
        if is_mismatch:
            mismatch_count += 1
            
        runner = {
            'index': i,
            'id': str(row1.get('registration_id')).strip() if row1.get('registration_id') else f'REG-{i}',
            'name': str(row1.get('name')).strip() if row1.get('name') else 'Unknown Runner',
            'email': str(row1.get('email')).strip() if row1.get('email') else '',
            'requestedLineup': req_lineup,
            'claimedRaceType': str(row1.get('claimed_race_type')).strip() if row1.get('claimed_race_type') else None,
            'claimedFinishTime': str(row1.get('claimed_finish_time')).strip() if row1.get('claimed_finish_time') else None,
            'verificationStatus': str(row1.get('verification_status')).strip() if row1.get('verification_status') else 'Unverified',
            'verificationSource': str(row1.get('verification_source')).strip() if row1.get('verification_source') else None,
            'verifiedRaceType': str(row1.get('verified_race_type')).strip() if row1.get('verified_race_type') else None,
            'verifiedFinishTime': str(row1.get('verified_finish_time')).strip() if row1.get('verified_finish_time') else None,
            'expectedLineup': exp_lineup,
            'confidenceScore': row1.get('confidence_score'),
            'decidedBy': row1.get('decided_by'),
            'decidedOn': row1.get('decided_on'),
            'remarks': str(row1.get('remarks')).strip() if row1.get('remarks') else None,
            'evidenceProvided': str(row2.get('evidence_provided')).strip() if row2.get('evidence_provided') else 'None',
            'resultLinkRaw': raw_text,
            'resultLinkClean': clean_url,
            'certificateFile': cert_path,
            'certificateType': cert_type,
            'isMismatch': is_mismatch
        }
        runners.append(runner)
        
    out_dir = os.path.join(base_dir, 'src', 'data')
    os.makedirs(out_dir, exist_ok=True)
    out_file = os.path.join(out_dir, 'runners.json')
    
    with open(out_file, 'w', encoding='utf-8') as f:
        json.dump(runners, f, indent=2, ensure_ascii=False)
        
    print(f'Successfully merged {len(runners)} records into {out_file}')
    print(f'- Mismatch records (Requested != Expected): {mismatch_count}')
    print(f'- Clean URLs extracted: {clean_url_count}')
    print(f'- Certificates mapped: {cert_count}')

if __name__ == '__main__':
    main()
